import pandas as pd
import utils.config as cfg
import utils.questions as q
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Side, Border, Alignment
from bs4 import BeautifulSoup
from tqdm import tqdm
import os
import pyminizip
import string
import random


class Cache:

    def __init__(self):
        self.formelfragen_pool = pd.read_excel(cfg.formelfragen_database_path)
        self.formelfragen_pool.index = self.formelfragen_pool["question_title"]

        self.single_choice_pool = pd.read_excel(cfg.single_choice_database_path)
        self.single_choice_pool.index = self.single_choice_pool["question_title"]

        self.bonuspunkte = {}
        for _, bonus in pd.read_excel(cfg.bonuspunkte_path).iterrows():
            if not pd.isna(bonus.Matrikelnummer):
                self.bonuspunkte[str(int(bonus.Matrikelnummer))] = bonus.Summe

    def get_bonuspoints(self, matrikelnummer):
        if pd.isna(matrikelnummer):
            return 0
        if str(int(matrikelnummer)) in list(self.bonuspunkte.keys()):
            return self.bonuspunkte[str(int(matrikelnummer))]
        else:
            return 0


class IliasReader:

    def __init__(self, path_to_export):
        self.ilias_overview = pd.read_excel(path_to_export, header=None)
        self.ilias_export = pd.ExcelFile(path_to_export)
        self.member = self._extract_members()

    def _extract_members(self):
        members = {}
        lines_with_students = ~pd.isna(self.ilias_overview.iloc[1:, 0])
        sheet_id = 1
        for idx, line in lines_with_students.items():
            if line:
                member_overview = self.ilias_overview.iloc[idx]
                member_overview.index = self._get_column_names(self.ilias_overview.iloc[idx - 1])
                member_detailed = self.ilias_export.parse(sheet_name=member_overview.loc["Name"][:cfg.max_excel_sheet_characters])

                members[member_overview.loc["Matrikelnummer"]] = Member(member_overview, member_detailed)
                sheet_id += 1
        return members

    def _get_column_names(self, row):
        ### TODO: explicit indexing ###
        identifier_column_names = list(self.ilias_overview.iloc[0, :cfg.num_ilias_overview_general_columns])
        question_column_names = list(row[cfg.num_ilias_overview_general_columns:-1])
        return identifier_column_names + question_column_names + ["Matrikelnummer"]


class QuestionParser:

    def __init__(self, overview, detailed_ilias):
        self.answers = detailed_ilias[detailed_ilias.iloc[:, 0].isin(cfg.question_identifiers)].copy()
        self.answers.columns = ["Type", "ID"]
        self.answers.loc[:, "idx_from"] = list(self.answers.index)
        self.answers.loc[:, "idx_to"] = list(self.answers.index)[1:] + [len(detailed_ilias)]
        self.answers.reset_index(inplace=True, drop=True)
        self.pool = self._parse_ilias_questions(overview, detailed_ilias)

    def _parse_ilias_questions(self, overview, detailed_ilias):
        question_pool = {}
        for idx, answer in self.answers.iterrows():
            if answer.Type == "Formelfrage":
                correct_ilias = overview[answer.ID]
                if cfg.use_exceptions:
                    question_pool[answer.ID] = q.FormelfrageExceptions(answer, detailed_ilias, correct_ilias)
                else:
                    question_pool[answer.ID] = q.Formelfrage(answer, detailed_ilias, correct_ilias)
            elif answer.Type == "Single Choice":
                correct_ilias = overview[answer.ID]
                question_pool[answer.ID] = q.SingleChoice(answer, detailed_ilias, correct_ilias)
            else:
                ### Placeholder for Freitext ###
                correct_ilias = overview[answer.ID]
                question_pool[answer.ID] = q.Frage(answer, detailed_ilias, correct_ilias)
        return question_pool


class Member:

    def __init__(self, overview, detailed):
        self.id = overview.loc["Matrikelnummer"]
        self._overview = overview
        self._detailed = detailed
        self.questions = QuestionParser(overview, detailed)
        self.bonus_points = cfg.cache.get_bonuspoints(self.id)
        self.name = self._overview.loc["Name"]

    @property
    def exam_points(self):
        points = 0
        for id, question in self.questions.pool.items():
            if question.correct:
                points += 1
        return points

    @property
    def points(self):
        return self.bonus_points + self.exam_points

    @property
    def grade(self):
        grading_helper = self.points / cfg.max_points * 100 >= cfg.grading_scheme
        return grading_helper.index[grading_helper][-1]


class ExportReview:

    def __init__(self, member_dict):
        self.member_export = pd.read_excel(cfg.ilias_member_export, index_col="Matrikelnummer")
        self.create_reviews(member_dict)
        self.create_psso_list(member_dict)
        self.zip_all_results()


    @staticmethod
    def set_cell_properties(cell, value, border, fill=None, wrap_text=False):
        cell.value = value
        cell.border = border
        cell.fill = fill
        cell.alignment = Alignment(wrap_text=wrap_text)

    def create_reviews(self, member_dict):
        print("Creating detailed reviews for Students...")
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        darker_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        brighter_fill = PatternFill(start_color="ffdbb6", end_color="ffdbb6", fill_type="solid")
        no_fill = PatternFill()
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))
        header = ["A#_Student ist ihre getaetigte Antwort",
                  "A#_Musterloesung ist die Richtige Antwort",
                  "A#_Punkte sind die resultierenden Punkte aus der jeweilgen Aufgabe",
                  "Die Bestehensgrenze betraegt für alle Teilnehmer 21 Punkte",
                  ""]

        for mtk_nr, member in tqdm(member_dict.items()):
            filename = f"{int(mtk_nr)}.xlsx"

            wb = Workbook()
            ws = wb.active

            # create header
            for idx, line in enumerate(header, 1):
                cell = ws[f"A{idx}"]
                cell.value = line
                cell.fill = yellow_fill
                cell = ws[f"B{idx}"]
                cell.fill = yellow_fill

            # create general information
            contents = ["Matrikelnummer", str(member.id),
                        "Name", member.name,
                        "Note", member.grade,
                        "Punkte in Klausur", str(int(member.exam_points)),
                        "Bonuspunkte", str(int(member.bonus_points)),
                        "Punkte Gesamt", str(int(member.points))]

            idx += 1
            for i in range(0, len(contents), 2):
                for j, col in enumerate(['A', 'B']):
                    cell = ws[f"{col}{idx}"]
                    cell.value = contents[i + j]
                    cell.border = thin_border
                idx += 1

            # Block with results per question
            idx += 1
            for fragennummer, frage in enumerate(member.questions.pool.values()):
                if frage._answer.Type == "Freitext eingeben":
                    continue

                # Aufgabenstellung
                plain_text = frage._question_description_main
                if frage._answer.Type == "Formelfrage":
                    soup = BeautifulSoup(plain_text, 'html.parser')
                    plain_text = soup.get_text('')
                    variables = frage._user_entry.iloc[:, 0]
                    values = frage._user_entry.iloc[:, 1]
                    for variable, value in zip(variables, values):
                        plain_text = plain_text.replace(str(variable), str(value))

                rows = [(f"A{fragennummer}_Aufgabenstellung", plain_text, no_fill, True),
                        (f"A{fragennummer}_Student", str(frage.result_user), darker_fill, False),
                        (f"A{fragennummer}_Musterloesung", str(frage.result_tool), darker_fill, False),
                        (f"A{fragennummer}_Punkte", str(int(frage.correct)), brighter_fill, False)]

                for label, value, fill, wrap_text in rows:
                    self.set_cell_properties(ws[f"A{idx}"], label, thin_border, fill, wrap_text)
                    self.set_cell_properties(ws[f"B{idx}"], value, thin_border, fill, wrap_text)
                    idx += 1

            # Adjust the column width to fit the content
            for col_idx, column in enumerate(ws.columns):
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                if col_idx == 0:
                    adjusted_width = max_length
                else:
                    adjusted_width = max_length / 8
                ws.column_dimensions[column[0].column_letter].width = adjusted_width

            if not os.path.exists(cfg.current_exam_path + "exports/students"):
                os.makedirs(cfg.current_exam_path + "exports/students")
            wb.save(f"{cfg.current_exam_path}exports/students/{filename}")

    def create_psso_list(self, member_dict):
        print("Creating PSSO list...")
        psso = {}
        for mtk_nr, member in tqdm(member_dict.items()):
            psso[mtk_nr] = {"Matrikelnummer": member.id,
                            "Name": member.name,
                            "Punkte in Klausur": member.exam_points,
                            "Bonuspunkte": member.bonus_points,
                            "Gesamtpunkte": member.points,
                            "Note": member.grade}
        self.psso_export = pd.DataFrame(psso).T
        grade_performance = self.psso_export["Note"].value_counts().sort_index()
        print(grade_performance)
        if not os.path.exists(cfg.current_exam_path + "exports/psso"):
            os.makedirs(cfg.current_exam_path + "exports/psso")
        self.psso_export.to_excel(cfg.current_exam_path + "exports/psso/psso_export.xlsx", index=False)
        self.psso_export.drop("Name", axis=1).to_excel(cfg.current_exam_path + "exports/psso/psso_export_anon.xlsx", index=False)

    @staticmethod
    def zip_and_protect(excel_file, zip_file, password):
        pyminizip.compress(excel_file, "", zip_file, password, 1)

    @staticmethod
    def generate_password(length=12, use_digits=True, use_special_chars=False):
        charset = string.ascii_letters
        if use_digits:
            charset += string.digits
        if use_special_chars:
            charset += string.punctuation
        password = ''.join(random.choice(charset) for _ in range(length))
        return password

    def zip_all_results(self):
        file_path = f"{cfg.current_exam_path}exports/students/"
        if not os.path.exists(cfg.current_exam_path + "exports/zip"):
            os.makedirs(cfg.current_exam_path + "exports/zip")

        key_chain = {}
        for file in os.listdir(file_path):
            mtk_nr = file[:-5]
            pw = self.generate_password()
            excel_file_path = f"{cfg.current_exam_path}exports/students/{file}"
            zip_file_path = f"{cfg.current_exam_path}exports/zip/{mtk_nr}.zip"
            self.zip_and_protect(excel_file_path, zip_file_path, pw)
            key_chain[mtk_nr] = {"E-Mail": self.member_export.loc[int(mtk_nr)]["E-Mail"],
                                 "Passwort": pw}

        self.key_chain = pd.DataFrame(key_chain).T
        self.key_chain.to_csv(cfg.current_exam_path + "exports/psso/mail_password.csv")
